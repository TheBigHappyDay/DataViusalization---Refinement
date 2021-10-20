from sys import prefix
import wx
from typing import TYPE_CHECKING, final
from openpyxl import workbook, worksheet
import matplotlib.pyplot as plt
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

file = "./Crash_Statistics_Victoria.xlsx"
class MyFrame(wx.Frame):
    def __init__(self, parent, title):
        super(MyFrame, self).__init__(parent, title =title, size = (600,200))
        self.panel = MyPanel(self)
 
class MyPanel(wx.Panel):
    def __init__(self, parent):
        print("---Program---")
        print("Data Loading...")
        super(MyPanel, self).__init__(parent)
        self.SetBackgroundColour('THISTLE')
        self.dateRange = []

        #Headings
        logoFont = wx.Font(wx.FontInfo(35).FaceName('Helvetica').Bold(bold=True))
        self.logo = wx.StaticText(self, -1, 'VL', pos = (5,0))
        self.logo.SetFont(logoFont)
        self.logo.SetForegroundColour('white')
        subHeadingFont = wx.Font(wx.FontInfo(10).FaceName('Helvetica').Bold(bold=True))
        self.subHeading = wx.StaticText(self,-1, 'Select Data Types:', (100,32))
        self.subHeading.SetFont(subHeadingFont)
        self.subHeading.SetForegroundColour('white')
        dateFont = wx.Font(wx.FontInfo(10).FaceName('Helvetica').Bold(bold=True))
        self.dateText = wx.StaticText(self,-1, 'Date From:', (300,5))
        self.dateText.SetFont(dateFont)
        self.dateText.SetForegroundColour('white')
        self.dateText = wx.StaticText(self,-1, 'Date To:', (430,5))
        self.dateText.SetFont(dateFont)
        self.dateText.SetForegroundColour('white')
        
        nameArray = []
        selectArray = []
        indexNum1 = 0
        indexNum2 = 0

        self.indexNum1 = indexNum1
        self.indexNum2 = indexNum2
        self.nameArray = nameArray
        self.selectArray = selectArray

        wb = load_workbook(file)
        ws = wb.active
        
        #FILLING (colNameArray) OF INDEX NAMES
        for row in range(1, 2):
            for col in range(1, 64):
                char = get_column_letter(col)
                info = ws[char + str(row)].value
                self.nameArray.append(info)

        #User Imput
        self.combobox1 = wx.ComboBox(self, choices = self.nameArray, pos = (100,50))
        self.Bind(wx.EVT_COMBOBOX, self.onCombo)
        self.combobox2 = wx.ComboBox(self, choices = self.nameArray, pos = (100,85))
        self.Bind(wx.EVT_COMBOBOX, self.onCombo)

        #Buttons
        self.generate = wx.Button(self, label = 'Generate', pos = (7,60), size = (60,35))
        self.generate.Bind(wx.EVT_BUTTON, self.onButton)
        self.dateFromButton = wx.Button(self, label = 'Enter Year', pos = (300,22), size = (120,30))
        self.dateFromButton.Bind(wx.EVT_BUTTON, self.dateFromEvent)
        self.dateToButton = wx.Button(self, label = 'Enter Year', pos = (430,22), size = (120,30))
        self.dateToButton.Bind(wx.EVT_BUTTON, self.dateToEvent)
        self.closeApp = wx.Button(self, label = 'Close',pos = (7,100),size = (50,25))
        self.closeApp.Bind(wx.EVT_BUTTON, self.closeWindow)

    #Events

    def onCombo(self, event):
        state = event.GetEventObject().GetValue()
        self.selectArray.append(state)
        xName = self.selectArray[0]
        yName = self.selectArray[1]
        
        self.xName = xName
        self.yName = yName

        x = self.nameArray.index(xName)
        y = self.nameArray.index(yName)

        self.x = x
        self.y = y

    def onButton(self, event):
        dataArray = []
        prefixArray = []

        wb = load_workbook(file)
        ws = wb.active

        self.x += 1
        self.y += 1

        for row in ws.iter_rows(min_row=2, min_col=self.x, max_row=74909, max_col=self.x):
            for cell in row:
                data = cell.value
                name = cell.coordinate
                name = list(name)
                if name[0].isalpha() and name[1].isalpha():
                    name.pop(0)
                    name[0] = 'E'
                else:
                    name[0] = 'E'
                dateCoord = ''.join(name)
                cellDate = ws[dateCoord].value
                cellDate = cellDate[-4:]
                cellDate = ''.join(cellDate)
                if cellDate in self.dateRange:
                    dataArray.append(data)

        for row in ws.iter_rows(min_row=2, min_col=self.y, max_row=74909, max_col=self.y):
            for cell in row:
                data = cell.value
                name = cell.coordinate
                name = list(name)
                if name[0].isalpha() and name[1].isalpha():
                    name.pop(0)
                    name[0] = 'E'
                else:
                    name[0] = 'E'
                dateCoord = ''.join(name)
                dateCoord = ''.join(name)
                cellDate = ws[dateCoord].value
                cellDate = cellDate[-4:]
                cellDate = ''.join(cellDate)
                if cellDate in self.dateRange:
                    prefixArray.append(data)
            
        plt.bar(prefixArray, dataArray)
        plt.grid(color = "green", linestyle = "--", linewidth = .5)
        plt.ylabel(self.xName)
        plt.xlabel(self.yName)
        plt.show()

    def dateFromEvent(self,event):
        dateFromBox = wx.TextEntryDialog(None, 'Date from?', 'Visual Labs', 'Enter Year')
        if dateFromBox.ShowModal() == wx.ID_OK:
            dateFromData = dateFromBox.GetValue()
        event.GetEventObject().SetLabel(dateFromData)
        self.startFrom = int(dateFromData)
        
    def dateToEvent(self,event):
        dateToBox = wx.TextEntryDialog(None, 'Date To?', 'Visual Labs', 'Enter Year')
        if dateToBox.ShowModal() == wx.ID_OK:
            dateToData = dateToBox.GetValue()
        event.GetEventObject().SetLabel(dateToData)
        endTo = int(dateToData)
        between = endTo - self.startFrom
  
        for i in range(between + 1):
            self.dateRange.append(str(self.startFrom))
            self.startFrom = self.startFrom + 1

    def closeWindow(self,event):
        quit()  


class MyApp(wx.App):
    def OnInit(self):
        self.frame = MyFrame(parent=None, title="Visual Labz Visualisation Tool©")
        self.frame.Show()
        return True
 
app = MyApp()
app.MainLoop()